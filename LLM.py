import requests
from fpdf import FPDF
from datetime import datetime, timedelta
import random
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.chart import BarChart, Reference
import markdown
import json 
from difflib import SequenceMatcher
import re
from functools import lru_cache
import openai
from dotenv import load_dotenv
import os
import tiktoken

load_dotenv()

class LLMResponder:
    def __init__(self, openai_api_key=None, index=None):
        self.context = []
        self.color_palette = [
            'FF9AA2', 'FFB7B2', 'FFDAC1', 'E2F0CB', 
            'B5EAD7', 'C7CEEA', 'F8B195', 'F67280',
            'C06C84', '6C5B7B', '355C7D'
        ]
        
        # Configuration GPT uniquement
        self.openai_api_key = openai_api_key or os.getenv('OPENAI_API_KEY')
        if not self.openai_api_key:
            raise ValueError("Clé API OpenAI requise")
            
        self.client = openai.OpenAI(api_key=self.openai_api_key)
        self.index = index
        self.tokenizer = tiktoken.get_encoding("cl100k_base")
        
        self._setup_darija_system()

    def _setup_darija_system(self):
        """Configure le système pour répondre en darija marocaine"""
        
        # Instructions pour GPT en darija
        self.SYSTEM_PROMPTS = {
            "darija_arabic": """أنت القطب الرقمي للفلاحة، مساعد ذكي متخصص في المجال الفلاحي المغربي.
            جاوب بالدارجة المغربية بحروف عربية، كن مهني ومفهوم.
            جاوب غير من الوثائق بلا ما تزيد.""",
            
            "darija_latin": """Nta Pole Digital D'Agriculture, assistant IA specialist f domaine dyal l fla7a l maghribiya.
            Jaweb b darija maghribiya b horouf latinia, kun professional o accessible.
            jawb ghi mn les fichiers li 3ndek matzid walo.""",
            
            "french": """Vous êtes le Pôle Digital D'Agriculture, assistant IA spécialisé dans l'agriculture marocaine.
            Répondez en français de manière professionnelle mais accessible.
            Utilisez juste les fichier que je te offre et non plus."""
        }
        # System prompt pour la traduction
        self.TRANSLATION_PROMPT = """Tu es un expert en traduction darija marocaine vers français.
        Traduis UNIQUEMENT la requête en français standard, sans ajouter d'explications.
        Concentre-toi sur les termes techniques agricoles et administratifs.
        Réponds seulement avec la traduction française."""
       
        # Vocabulaire étendu pour la détection de darija
        self.darija_latin_vocab = {
            # Salutations et politesse
            "salam", "ahlan", "salamu", "labas", "bikhir", "hamdulillah", "baraka", "bslama",
            # Questions et interrogation
            "ash", "ashno", "kifash", "kif", "fin", "mnin", "wqt", "waqtash", "3lash", "3la", 
            "wach", "wash", "aji", "ajibo", "guli", "gul", "quli", "qul",
            # Mots courants
            "bzaf", "bzzaf", "ktir", "shwiya", "chwiya", "qalil", "zwin", "zwina", "ghzal",
            "ghzala", "bikhir", "mlih", "mzyan", "khaib", "khayb", "ma3ruf", "m3ruf",
            # Verbes d'action agricole
            "zarb", "zr3", "zra3", "sqa", "sqaya", "7sad", "hsad", "qta3", "qla3", "shuf",
            # Agriculture en darija
            "fla7a", "flaha", "zar3", "zr3", "7qla", "hqla", "ma", "trab", "bhayem", "b7ayem",
            "dgag", "djaj", "bqra", "bqar", "khalul", "khlul", "ghnam", "ghanam", "7lib", "hlib",
            # Conjonctions et prépositions
            "w", "wa", "wla", "ola", "la", "li", "lli", "dyal", "dial", "d", "f", "fi", "mn", "men",
            "3nd", "3and", "m3a", "ma3", "bla", "bila", "7ta", "hta", "ghir", "ghi", "kan", "kan",
        }
        
        self.arabic_patterns = [
            r'[\u0600-\u06FF]',  # Caractères arabes de base
            r'[\u0750-\u077F]',  # Suppléments arabes
            r'[\uFB50-\uFDFF]',  # Formes de présentation arabes A
            r'[\uFE70-\uFEFF]'   # Formes de présentation arabes B
        ]

    def contient_arabe(self, texte):
        """Détecte si le texte contient des caractères arabes"""
        return bool(re.search(r'[\u0600-\u06FF]', texte))

    def is_greeting(self, query):
        """Détecte si la requête est une salutation"""
        query_clean = query.lower().strip()
        
        # Salutations exactes
        exact_greetings = [
            "salam", "ahlan", "مرحبا", "السلام", "السلام عليكم",
            "bonjour", "salut", "hello", "hi", "coucou", "hey"
        ]
        
        # Vérification exacte
        if query_clean in exact_greetings:
            return True
            
        # Vérification avec des variantes
        greeting_patterns = [
            r'^salam\s*(aleikum|3likoum)?$',
            r'^(bonjour|salut)\s*!?$',
            r'^(hello|hi|hey)\s*!?$',
            r'^\s*مرحبا\s*$',
            r'^\s*السلام\s*(عليكم)?\s*$'
        ]
        
        for pattern in greeting_patterns:
            if re.match(pattern, query_clean, re.IGNORECASE):
                return True
                
        return False

    def detect_language_advanced(self, query):
        """Détection avancée de la langue avec support des 3 langues"""
        query_lower = query.lower().strip()
        
        # Détection de l'arabe
        if self.contient_arabe(query):
            return "darija_arabic"
        
        # Salutations spécifiques
        if query_lower in ["salam", "ahlan"]:
            return "darija_latin"
            
        # Détection de mots darija en lettres latines
        darija_latin_count = sum(1 for word in self.darija_latin_vocab if word in query_lower)
        
        # Si plus de 2 mots darija détectés, c'est probablement de la darija
        if darija_latin_count >= 2:
            return "darija_latin"
        
        # Détection de mots darija courants
        common_darija = ["kifash", "wach", "wash", "bzaf", "chwiya", "dyal", "3la", "mn", "fin"]
        if any(word in query_lower for word in common_darija):
            return "darija_latin"
        
        return "french"

    """ def translate_query_to_french(self, query, detected_lang):
        
        
        # Si c'est déjà en français, pas besoin de traduire
        if detected_lang == "french":
            print(f"[TRANSLATION] Requête déjà en français: {query}")
            return query
        
        try:
            print(f"[TRANSLATION] Traduction de '{query}' ({detected_lang}) vers français...")
            
            response = self.client.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": self.TRANSLATION_PROMPT},
                    {"role": "user", "content": f"Traduis cette requête en français: {query}"}
                ],
                temperature=0.1,  # Très bas pour traduction précise
                max_tokens=200,
                top_p=0.9
            )
            
            translated = response.choices[0].message.content.strip()
            print(f"[TRANSLATION] Résultat: '{translated}'")
            
            return translated 
            
        except Exception as e:
            print(f"[TRANSLATION] Erreur lors de la traduction: {e}")
            # Fallback: retourner la requête originale
            return query """

    def get_greeting_by_language(self, language):
        """Retourne une salutation appropriée selon la langue détectée"""
        greetings = {
            "darija_arabic": "السلام عليكم! أنا القطب الرقمي للفلاحة، مساعدك في المجال الفلاحي. كيف يمكنني مساعدتك اليوم؟",
            "darija_latin": "Salam! Ana Pole Digital D'Agriculture, mosa3id dyalek f domaine dyal l fla7a. Kifash n9der n3awnek lyoum?",
            "french": "Bonjour ! Je suis le Pôle Digital D'Agriculture, votre assistant spécialisé dans le domaine agricole. Comment puis-je vous aider aujourd'hui ?"
        }
        return greetings.get(language, greetings["french"])

    def update_context(self, query, response, max_context_length=5):
        """Met à jour le contexte avec la nouvelle question et réponse."""
        if len(self.context) >= max_context_length:
            self.context.pop(0)
        self.context.append({"query": query, "response": response})

    def get_context(self):
        """Retourne le contexte sous forme de chaîne de caractères."""
        return "\n".join(f"Question: {entry['query']}\nRéponse: {entry['response']}" for entry in self.context)

    def ask_gpt_darija(self, question, detected_lang="darija_latin", retrieved_context=None):
        """Fonction principale pour interroger GPT en darija"""
        
        # Préparation du système prompt
        system_prompt = self.SYSTEM_PROMPTS.get(detected_lang, self.SYSTEM_PROMPTS["darija_latin"])
        
        # Construction du prompt utilisateur
        user_prompt = self._build_darija_prompt(question, detected_lang, retrieved_context)
        
        # Monitoring des tokens
        total_tokens = len(self.tokenizer.encode(system_prompt + user_prompt))
        print(f"🔢 Tokens utilisés : {total_tokens}")
        
        try:
            response = self.client.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.3,
                max_tokens=1500,
                top_p=0.9,
                frequency_penalty=0.2,
                presence_penalty=0.2
            )
            
            return response.choices[0].message.content
            
        except Exception as e:
            print(f"Erreur GPT: {e}")
            return f"Sma7 lia, kan 3andi mushkil f jawab. 3awd t9ad men ba3d. (Erreur: {str(e)})"

    def _build_darija_prompt(self, question, detected_lang, retrieved_context=None):
        """Construit le prompt pour GPT en darija"""
        
        context_part = ""
        if self.context:
            context_part = f"\n### Historique de la conversation ###\n{self.get_context()}\n"
        
        if retrieved_context:
            context_part += f"\n### Documents de référence ###\n{retrieved_context}\n"
        
        # Instructions selon la langue
        if detected_lang == "darija_arabic":
            instructions = """
            ### تعليمات ###
            - جاوب بالدارجة المغربية بحروف عربية
            - كن مهني ومفهوم
            - غير من الوثائق بلا ما تزيد
            - ركز على السؤال المطروح فقط
            - إذا وجدت معلومات في الوثائق المرجعية، استخدمها في إجابتك
            """
        elif detected_lang == "darija_latin":
            instructions = """
            ### Ta3limat ###
            - Jaweb b darija maghribiya b horouf latinia
            - Kun professional o mafhum
            - Ste3mel ghi lwata2i9 bla matzid
            - Rkez 3la su2al li t9ad ghir
            - Ila l9iti ma3lomat f documents, ste3melhom f jawab dyalek
            """
        else:  # french
            instructions = """
            ### Instructions ###
            - Répondez en français
            - Soyez professionnel et accessible
            - Utilisez juste les fichier
            - Concentrez-vous uniquement sur la question posée
            - Si vous trouvez des informations dans les documents de référence, utilisez-les dans votre réponse
            """
        
        return f"""{instructions}
            {context_part}
            ### Su2al / Question ###
            {question}

            ### Jawab / Réponse ###"""

    def generate_response(self, query, retrieved_answer=None):
        """Génère une réponse en darija via GPT"""
        
        print(f"[DEBUG] Requête reçue: '{query}'")
        
        # Détection prioritaire des salutations
        if self.is_greeting(query):
            print("[DEBUG] Salutation détectée!")
            detected_lang = self.detect_language_advanced(query)
            print(f"[DEBUG] Langue détectée pour salutation: {detected_lang}")
            greeting_response = self.get_greeting_by_language(detected_lang)
            
            # Mise à jour du contexte
            self.update_context(query, greeting_response)
            
            yield f"data: {json.dumps({'content': greeting_response, 'finished': True, 'language': detected_lang})}\n\n"
            return

        # Détection de la langue
        detected_lang = self.detect_language_advanced(query)
        print(f"[DEBUG] Langue détectée: {detected_lang}")
        
        # Si pas de darija détectée, forcer darija latin
        if detected_lang == "french":
            detected_lang = "darija_latin"
            print("[DEBUG] Langue forcée vers darija_latin")
        
        try:
            # Préparation du contexte des documents
            context_docs = None
            if retrieved_answer:
                if isinstance(retrieved_answer, list):
                    context_docs = "\n\n".join(retrieved_answer[:3])  # Limiter à 3 documents
                else:
                    context_docs = str(retrieved_answer)
                
                print(f"[DEBUG] Contexte documentaire trouvé: {len(context_docs) if context_docs else 0} caractères")
            else:
                print("[DEBUG] Aucun contexte documentaire fourni")
            
            # Appel à GPT avec la requête ORIGINALE pour garder la réponse en darija
            response_text = self.ask_gpt_darija(query, detected_lang, context_docs)
            
            # Mise à jour du contexte
            self.update_context(query, response_text)
            
            # Formatage HTML
            formatted_response = self._format_response(response_text)
            
            yield f"data: {json.dumps({'content': formatted_response, 'finished': True, 'is_html': True, 'language': detected_lang})}\n\n"
            
        except Exception as e:
            error_msg = f"Sma7 lia, kan 3andi mushkil. 3awd t9ad men ba3d. (Erreur: {str(e)})"
            yield f"data: {json.dumps({'error': error_msg, 'finished': True, 'language': 'darija_latin'})}\n\n"

    def _format_response(self, text):
        """Formate la réponse en HTML via Markdown"""
        html = markdown.markdown(text)
        return html

    def save_as_pdf(self, query, response_text):
        """Sauvegarde la question et la réponse en PDF."""
        pdf = FPDF()
        pdf.add_page()
        pdf.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
        pdf.set_font('DejaVu', '', 12)
        pdf.cell(200, 10, txt="Question:",  ln=True) # type: ignore
        pdf.multi_cell(0, 10, txt=query)# type: ignore
        pdf.cell(200, 10, txt="Réponse:", ln=True)# type: ignore
        pdf.multi_cell(0, 10, txt=response_text)# type: ignore
        pdf.output("response.pdf")
        print("Réponse enregistrée en PDF.")

    def save_as_text(self, query, response_text):
        """Sauvegarde la question et la réponse en texte."""
        with open("response.txt", 'w', encoding='utf-8') as file:
            file.write(f"Question:\n{query}\n\nRéponse:\n{response_text}")
        print("Réponse enregistrée en texte.")